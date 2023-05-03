from openpyxl import load_workbook, Workbook


result = Workbook()  # Таблица результата
result_sheets = result.active
result_sheets.title = 'Результаты'
result_sheets.append(['#', 'id Обьявления','Конверсии/Заявка','Конверсии/Переход в Ватсап','Конверсии/Клик по номеру телефона'])

wb = load_workbook('./Директ.xlsx')  # Таблица 1
ws = wb.active
columns_a = ws['A']  # Дата
columns_f = ws['F']  # № Объявления
columns_t = ws['T']  # Конверсии/Заявка
columns_u = ws['U']  # Конверсии/Переход в Ватсап
columns_v = ws['V']  # Конверсии/Клик по номеру телефона


wb_2 = load_workbook('./Заявки.xlsx')  #Таблица 2
ws_2 = wb_2.active
columns_a_num = ws_2['A']  # Номер "#"
columns_b = ws_2['B']  # Время
columns_g = ws_2['G']  # Обьявление id

dict_bboard_id = dict()
dict_bboard_id_all_info = {}
for a, b, g in zip(columns_a_num, columns_b, columns_g):  # Создание словаря по таблице 2 c {"id/#": дата}
    if b.value:
        dict_bboard_id_all_info.setdefault(f'{g.value}/{a.value}', []).append('.'.join(str(b.value).split(' ')[0].split('-')[::-1]))


for val in dict_bboard_id_all_info.keys():
    date = "".join(dict_bboard_id_all_info[val])

    for a, f, t, u, v in zip(columns_a, columns_f, columns_t,columns_u, columns_v):
        if val.split('/')[0] == str(f.value) and a.value == date:
            print(val.split('/')[1], f.value, t.value, u.value, v.value)
            result_sheets.append([val.split('/')[1], f.value, t.value, u.value, v.value])


result.save('result.xlsx')  # Запись в таблицу результата
